import os
import time
import openpyxl

# Folder path -- RPA trigger files
TRIGGER_FOLDER_PATH = "D:\\WorkSpace\\Python\\rpa\\"
# Excel file -- RPA && trigger file name
RPA_FILENAME_DICT_PATH = "D:\\WorkSpace\\Python\\rpa\\rpa_filename.xlsx"

# read Excel file
def read_excel():
    dict = {}
    wb = openpyxl.load_workbook(RPA_FILENAME_DICT_PATH)
    sheet = wb["Sheet1"]
    for i in range(sheet.min_row + 1, sheet.max_row + 1):
        rpa = sheet.cell(i, 1).value
        filename = sheet.cell(i, 2).value
        dict[rpa] = filename
    print(dict)
    return dict

# create file trigger rpa
def create_file(filename):
    if not os.path.exists(TRIGGER_FOLDER_PATH):
        print("folder is not exist")
    open(TRIGGER_FOLDER_PATH + filename, 'w').close()
    print(f"create file -- {filename}, trigger RPA!")

# wait file delete
def wait_delete_file(filename):
    filename = TRIGGER_FOLDER_PATH + filename
    while os.path.exists(filename):
        time.sleep(2)
        print("wait RPA 2s!")
    print("finish waiting RPA!")

RF_DICT = read_excel()